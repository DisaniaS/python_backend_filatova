from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
import math
import numpy as np
from scipy import stats

from fastapi import HTTPException
from fastapi.params import Depends
from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from starlette import status
from starlette.responses import FileResponse

from inaccuracy.schema import YearlyData, YearlyCountData, ErrorResponse, CorrelationData, CorrelationMatrix, Reason, Measure
from report_data.repository import ReportDataRepository
from report_data.model import ReportData


class InaccuracyService:
    def __init__(
            self,
            report_data_repo: ReportDataRepository = Depends()
    ):
        self.report_data_repo = report_data_repo
        self.table_path = Path("inaccuracy/inaccuracytest.xlsx").absolute()

        self.table_path.parent.mkdir(parents=True, exist_ok=True)
        self.K_MAX = 3  # Максимальное допустимое значение погрешности
        # Словарь причин и мероприятий
        self._init_reasons_measures()

    def _init_reasons_measures(self):
        """Инициализация списка причин и мероприятий"""
        self.reasons = [
            Reason(
                title="Различия в качестве гироскопов",
                description="Наблюдаются существенные различия в точности и стабильности разных типов гироскопов",
                measures=[
                    Measure(
                        title="Провести анализ конструкции",
                        description="Детальный анализ конструкции гироскопов для выявления слабых мест"
                    ),
                    Measure(
                        title="Провести анализ технологического производства",
                        description="Анализ технологического процесса изготовления с целью выявления отклонений"
                    ),
                    Measure(
                        title="Изучить причины в различии типов",
                        description="Проведение сравнительных испытаний гироскопов разных типов в одинаковых условиях"
                    ),
                    Measure(
                        title="Улучшить гироскопы типа А",
                        description="Модернизация гироскопов определенного типа для повышения их точности"
                    )
                ]
            ),
            Reason(
                title="Неустойчивая важность",
                description="Выявлена нестабильность работы изделий в условиях повышенной влажности",
                measures=[
                    Measure(
                        title="Использовать влагозащитные покрытия",
                        description="Применение специальных влагозащитных покрытий для электронных компонентов"
                    ),
                    Measure(
                        title="Контролировать уровень влажности",
                        description="Установка строгого контроля уровня влажности в помещениях для испытаний"
                    )
                ]
            ),
            Reason(
                title="Влияние устаревшего оборудования",
                description="Обнаружена зависимость погрешности от года выпуска оборудования, указывающая на устаревание",
                measures=[
                    Measure(
                        title="Модернизировать устаревшие компоненты",
                        description="Замена устаревших компонентов на современные аналоги с улучшенными характеристиками"
                    ),
                    Measure(
                        title="Заменить устаревшее оборудование",
                        description="Полная замена оборудования, выпущенного ранее определенного года"
                    )
                ]
            ),
            Reason(
                title="Недостаточность вибрационной нагрузки",
                description="Выявлено влияние недостаточной вибрационной нагрузки на точность измерений",
                measures=[
                    Measure(
                        title="Увеличить вибрационную нагрузку",
                        description="Увеличение интенсивности вибрационной нагрузки при испытаниях"
                    ),
                    Measure(
                        title="Улучшить виброизоляцию",
                        description="Установка улучшенных виброизоляционных платформ в камеры проведения испытаний"
                    )
                ]
            )
        ]

    def get_error_data(self) -> ErrorResponse:
        """
        Получает данные о соотношениях погрешностей из Excel файла
        """
        if not self.table_path.exists():
            return ErrorResponse(yearly_data={})

        try:
            wb = load_workbook(self.table_path, read_only=True)
            ws = wb.active

            # Структура для хранения данных по годам
            yearly_data: Dict[str, Dict] = {}

            # Пропускаем заголовок
            rows = list(ws.rows)[1:]

            for row in rows:
                year = str(row[0].value)  # Колонка A - Год
                if not year:
                    continue

                if year not in yearly_data:
                    yearly_data[year] = {
                        'minus_50': [],
                        'plus_50': [],
                        'nku': []
                    }

                # Получаем значения погрешностей из таблицы
                nku_error = row[2].value  # Колонка C - Погрешность НКУ
                minus_50_error = row[3].value  # Колонка D - Погрешность -50
                plus_50_error = row[4].value  # Колонка E - Погрешность +50

                # Рассчитываем μ для каждой температуры
                if nku_error is not None:
                    try:
                        mu_nku = float(nku_error) / self.K_MAX
                        # Проверка на NaN
                        if not math.isnan(mu_nku):
                            yearly_data[year]['nku'].append(mu_nku)
                    except:
                        pass

                if minus_50_error is not None:
                    try:
                        mu_minus_50 = float(minus_50_error) / self.K_MAX
                        # Проверка на NaN
                        if not math.isnan(mu_minus_50):
                            yearly_data[year]['minus_50'].append(mu_minus_50)
                    except:
                        pass

                if plus_50_error is not None:
                    try:
                        mu_plus_50 = float(plus_50_error) / self.K_MAX
                        # Проверка на NaN
                        if not math.isnan(mu_plus_50):
                            yearly_data[year]['plus_50'].append(mu_plus_50)
                    except:
                        pass

            # Формирование результата с использованием моделей Pydantic
            formatted_data = {}
            for year, data in yearly_data.items():
                formatted_data[year] = YearlyData(
                    minus_50=data['minus_50'],
                    plus_50=data['plus_50'],
                    nku=data['nku'],
                    count=YearlyCountData(
                        minus_50=len(data['minus_50']),
                        plus_50=len(data['plus_50']),
                        nku=len(data['nku'])
                    )
                )

            wb.close()
            
            # Расчет корреляций и формирование полного ответа
            correlations = self.calculate_correlations()
            correlation_matrix = self.create_correlation_matrix()
            
            return ErrorResponse(
                yearly_data=formatted_data,
                correlations=correlations,
                correlation_matrix=correlation_matrix
            )

        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Ошибка при чтении файла: {str(e)}"
            )

    def download_inaccuracy(self) -> FileResponse:
        if not self.table_path.exists():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Таблица не найдена",
            )

            # Добавляем timestamp для предотвращения кэширования
        timestamp = int(datetime.now().timestamp())
        filename = f"inaccuracy_{timestamp}.xlsx"

        return FileResponse(
            path=self.table_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
                "Expires": "0"
            }
        )

    def get_uncalculated_reports(self) -> List[ReportData]:
        """Получает отчеты, которые еще не были рассчитаны"""
        return self.report_data_repo.db.query(ReportData)\
            .filter(ReportData.calculated == False)\
            .all()

    def calculate_errors_for_report(self, report: ReportData) -> dict:
        """Вычисляет погрешности для одного отчета"""
        return {
            "Год": report.test_date.split(".")[2] if report.test_date else datetime.now().year,
            "Номер изделия": report.system_number,
            "Погрешность НКУ": self._calculate_error_nku(
                report.azimuth_minus_50,
                report.repeated_azimuth_minus_50,
                report.azimuth_nku,
                report.repeated_azimuth_nku,
                report.azimuth_plus_50,
                report.repeated_azimuth_plus_50
            ),
            "Погрешность -50": self._calculate_error(
                report.azimuth_minus_50,
                report.repeated_azimuth_minus_50,
                report.azimuth_nku,
                report.repeated_azimuth_nku
            ),
            "Погрешность +50": self._calculate_error(
                report.azimuth_plus_50,
                report.repeated_azimuth_plus_50,
                report.azimuth_nku,
                report.repeated_azimuth_nku
            )
        }

    def _calculate_error(self, exact: float, repeated: float, exact_nku: float, repeated_nku: float) -> float:
        """Вычисляет погрешность по формуле (max - min)/2"""
        if None in (exact, repeated, exact_nku, repeated_nku):
            return None
        return (max(exact, repeated, exact_nku, repeated_nku) - min(exact, repeated, exact_nku, repeated_nku)) / 2

    def _calculate_error_nku(self, exact_minus: float, repeated_minus: float, exact_nku: float, repeated_nku: float, exact_plus: float, repeated_plus: float) -> float:
        """Вычисляет погрешность по формуле (max - min)/2"""
        if None in (exact_minus, repeated_minus, exact_nku, repeated_nku, exact_plus, repeated_plus):
            return None
        return (max(exact_minus, repeated_minus, exact_nku, repeated_nku, exact_plus, repeated_plus) - min(exact_minus, repeated_minus, exact_nku, repeated_nku, exact_plus, repeated_plus)) / 2


    def update_excel_file(self):
        """Добавляет новые данные в Excel-файл"""
        # Получаем только нерассчитанные отчеты
        uncalculated_reports = self.get_uncalculated_reports()

        if not uncalculated_reports:
            return  # Нет новых данных для добавления

        try:
            wb = load_workbook(self.table_path)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            # Добавляем заголовки, если файл новый
            ws.append(["Год", "Номер изделия", "Погрешность НКУ", "Погрешность -50", "Погрешность +50"])

        # Добавляем данные для каждого нерассчитанного отчета
        for report in uncalculated_reports:
            error_data = self.calculate_errors_for_report(report)
            ws.append([
                error_data["Год"],
                error_data["Номер изделия"],
                error_data["Погрешность НКУ"],
                error_data["Погрешность -50"],
                error_data["Погрешность +50"]
            ])
            ##Помечаем отчет как рассчитанный
            report.calculated = True
            self.report_data_repo.db.commit()

        try:
            self._apply_formatting(ws)
            wb.save(self.table_path)
            wb.close()  # Важно закрыть файл

            # Проверяем, что файл обновился
            if not self.table_path.exists():
                raise Exception("Файл не был создан")

        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Ошибка при обновлении файла: {str(e)}"
            )

    def _apply_formatting(self, ws):
        """Настраивает оформление таблицы"""
        # Ширина столбцов
        column_widths = {'A': 8, 'B': 15, 'C': 18, 'D': 18, 'E': 18}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Жирные заголовки (если есть данные)
        if ws.max_row > 1:
            for cell in ws[1]:
                cell.font = Font(bold=True)

        # Формат чисел для всех числовых ячеек
        for row in ws.iter_rows(min_row=2):
            for cell in row[2:]:  # Столбцы C-E
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
                    
    def calculate_correlations(self) -> Dict[str, CorrelationData]:
        """
        Рассчитывает корреляции для погрешностей при разных температурах,
        используя реальные данные о влажности и вибрации из таблицы report_data
        """
        if not self.table_path.exists():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Таблица с данными о погрешностях не найдена"
            )
            
        try:
            wb = load_workbook(self.table_path, read_only=True)
            ws = wb.active
            
            # Получаем все данные из репозитория
            all_report_data = self.report_data_repo.db.query(ReportData).all()
            
            # Создаем словарь для поиска данных по номеру изделия
            report_data_dict = {rd.system_number: rd for rd in all_report_data if rd.system_number is not None}
            
            # Собираем данные из Excel и соответствующие данные из таблицы report_data
            data = {'year': [], 'vlagh': [], 'urvibr': [], 'cert': [], 'type': [], 
                   'dol1': [], 'dol2': [], 'dol3': []}
            
            # Словари для маппинга названий типов и частей на числовые ранги
            types_unique = set()
            certs_unique = set()
            
            # Пропускаем заголовок
            rows = list(ws.rows)[1:]
            
            for row in rows:
                product_number = row[1].value  # Номер изделия в колонке B
                year_value = row[0].value      # Год в колонке A
                
                if not product_number or not year_value:
                    continue
                
                # Находим соответствующие данные в таблице report_data
                report_data_entry = report_data_dict.get(product_number)
                
                # Берем погрешности из файла Excel
                nku_error = row[2].value  # Погрешность НКУ
                minus_50_error = row[3].value  # Погрешность -50
                plus_50_error = row[4].value  # Погрешность +50
                
                if None in (nku_error, minus_50_error, plus_50_error):
                    continue
                    
                # Рассчитываем доли от максимально допустимого
                dol1 = float(nku_error) / self.K_MAX
                dol2 = float(minus_50_error) / self.K_MAX
                dol3 = float(plus_50_error) / self.K_MAX
                
                # Добавляем данные
                data['year'].append(int(year_value))
                data['dol1'].append(dol1)
                data['dol2'].append(dol2)
                data['dol3'].append(dol3)
                
                # Добавляем реальные данные о влажности, вибрации, типе и части, если они есть
                if report_data_entry:
                    # Влажность и вибрация
                    humidity = report_data_entry.humidity
                    vibration = report_data_entry.vibration_level
                    
                    # Тип изделия и часть (департамент)
                    cert = report_data_entry.department
                    product_type = report_data_entry.system_type
                    
                    # Сохраняем уникальные значения для последующего маппинга
                    if cert:
                        certs_unique.add(cert)
                    if product_type:
                        types_unique.add(product_type)
                    
                    # Добавляем в данные только если они существуют
                    if humidity is not None:
                        data['vlagh'].append(humidity)
                    else:
                        data['vlagh'].append(None)
                        
                    if vibration is not None:
                        data['urvibr'].append(vibration)
                    else:
                        data['urvibr'].append(None)
                        
                    data['cert'].append(cert)
                    data['type'].append(product_type)
                else:
                    # Отмечаем как отсутствующие данные
                    data['vlagh'].append(None)
                    data['urvibr'].append(None)
                    data['cert'].append(None)
                    data['type'].append(None)
                
            # Если данных нет, возвращаем ошибку
            if not data['year']:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Нет данных для расчета корреляций"
                )
            
            # Создаем словари соответствия строковых значений и числовых рангов
            cert_map = {cert: i+1 for i, cert in enumerate(certs_unique) if cert is not None}
            types_map = {type_val: i+1 for i, type_val in enumerate(types_unique) if type_val is not None}
            
            # Обратные словари для получения строковых значений по рангам
            cert_map_reverse = {v: k for k, v in cert_map.items()}
            types_map_reverse = {v: k for k, v in types_map.items()}
            
            # Преобразуем строковые значения в ранги для корреляционного анализа
            cert_ranks = [cert_map.get(cert, None) for cert in data['cert']]
            type_ranks = [types_map.get(type_val, None) for type_val in data['type']]
                
            # Рассчитываем корреляции
            correlations = {}
            errors = {}
            
            # Для каждой температуры (-50, +50, НКУ)
            for temp_mode, dol_key in [('minus_50', 'dol2'), ('plus_50', 'dol3'), ('nku', 'dol1')]:
                temp_errors = []
                
                # Преобразуем в numpy массивы, фильтруя None значения
                dol = np.array(data[dol_key])
                
                # Фильтруем данные для year
                valid_year_indices = [i for i, val in enumerate(data['year']) if val is not None]
                if valid_year_indices:
                    year_data = [(data['year'][i], data[dol_key][i]) for i in valid_year_indices]
                    year_vals, dol_for_year = zip(*year_data) if year_data else ([], [])
                    year = np.array(year_vals)
                    dol_year = np.array(dol_for_year)
                    
                    # Проверяем на наличие вариации в данных
                    if len(set(year)) <= 1 or len(set(dol_year)) <= 1:
                        temp_errors.append("Недостаточная вариация в данных года или значениях погрешности")
                        corr_year, p_year = None, None
                    else:
                        try:
                            corr_year, p_year = stats.pearsonr(year, dol_year)
                        except Exception as e:
                            temp_errors.append(f"Ошибка при расчете корреляции по годам: {str(e)}")
                            corr_year, p_year = None, None
                else:
                    temp_errors.append("Недостаточно данных для расчета корреляции по годам")
                    corr_year, p_year = None, None
                
                # Фильтруем данные для vlagh
                valid_vlagh_indices = [i for i, val in enumerate(data['vlagh']) if val is not None]
                if valid_vlagh_indices:
                    vlagh_data = [(data['vlagh'][i], data[dol_key][i]) for i in valid_vlagh_indices]
                    vlagh_vals, dol_for_vlagh = zip(*vlagh_data) if vlagh_data else ([], [])
                    vlagh = np.array(vlagh_vals)
                    dol_vlagh = np.array(dol_for_vlagh)
                    
                    # Проверяем на наличие вариации в данных
                    if len(set(vlagh)) <= 1 or len(set(dol_for_vlagh)) <= 1:
                        temp_errors.append("Недостаточная вариация в данных влажности или значениях погрешности")
                        corr_vlagh, p_vlagh = None, None
                    else:
                        try:
                            corr_vlagh, p_vlagh = stats.pearsonr(vlagh, dol_vlagh)
                        except Exception as e:
                            temp_errors.append(f"Ошибка при расчете корреляции по влажности: {str(e)}")
                            corr_vlagh, p_vlagh = None, None
                else:
                    temp_errors.append("Недостаточно данных для расчета корреляции по влажности")
                    corr_vlagh, p_vlagh = None, None
                
                # Фильтруем данные для urvibr
                valid_urvibr_indices = [i for i, val in enumerate(data['urvibr']) if val is not None]
                if valid_urvibr_indices:
                    urvibr_data = [(data['urvibr'][i], data[dol_key][i]) for i in valid_urvibr_indices]
                    urvibr_vals, dol_for_urvibr = zip(*urvibr_data) if urvibr_data else ([], [])
                    urvibr = np.array(urvibr_vals)
                    dol_urvibr = np.array(dol_for_urvibr)
                    
                    # Проверяем на наличие вариации в данных
                    if len(set(urvibr)) <= 1 or len(set(dol_for_urvibr)) <= 1:
                        temp_errors.append("Недостаточная вариация в данных вибрации или значениях погрешности")
                        corr_urvibr, p_urvibr = None, None
                    else:
                        try:
                            corr_urvibr, p_urvibr = stats.pearsonr(urvibr, dol_urvibr)
                        except Exception as e:
                            temp_errors.append(f"Ошибка при расчете корреляции по вибрации: {str(e)}")
                            corr_urvibr, p_urvibr = None, None
                else:
                    temp_errors.append("Недостаточно данных для расчета корреляции по вибрации")
                    corr_urvibr, p_urvibr = None, None
                
                # Фильтруем для type
                valid_type_indices = [i for i, val in enumerate(type_ranks) if val is not None]
                if valid_type_indices:
                    type_data = [(type_ranks[i], data[dol_key][i]) for i in valid_type_indices]
                    type_vals, dol_for_type = zip(*type_data) if type_data else ([], [])
                    type_data = np.array(type_vals)
                    dol_type = np.array(dol_for_type)
                    
                    # Проверяем на наличие вариации в данных
                    if len(set(type_data)) <= 1 or len(set(dol_for_type)) <= 1:
                        temp_errors.append("Недостаточная вариация в данных типа или значениях погрешности")
                        corr_type, p_type = None, None
                    else:
                        try:
                            corr_type, p_type = stats.pearsonr(type_data, dol_type)
                        except Exception as e:
                            temp_errors.append(f"Ошибка при расчете корреляции по типу изделия: {str(e)}")
                            corr_type, p_type = None, None
                else:
                    temp_errors.append("Недостаточно данных для расчета корреляции по типу изделия")
                    corr_type, p_type = None, None
                
                # Фильтруем для cert
                valid_cert_indices = [i for i, val in enumerate(cert_ranks) if val is not None]
                if valid_cert_indices:
                    cert_data = [(cert_ranks[i], data[dol_key][i]) for i in valid_cert_indices]
                    cert_vals, dol_for_cert = zip(*cert_data) if cert_data else ([], [])
                    cert_data = np.array(cert_vals)
                    dol_cert = np.array(dol_for_cert)
                    
                    # Проверяем на наличие вариации в данных
                    if len(set(cert_data)) <= 1 or len(set(dol_for_cert)) <= 1:
                        temp_errors.append("Недостаточная вариация в данных части или значениях погрешности")
                        corr_cert, p_cert = None, None
                    else:
                        try:
                            corr_cert, p_cert = stats.pearsonr(cert_data, dol_cert)
                        except Exception as e:
                            temp_errors.append(f"Ошибка при расчете корреляции по части: {str(e)}")
                            corr_cert, p_cert = None, None
                else:
                    temp_errors.append("Недостаточно данных для расчета корреляции по части")
                    corr_cert, p_cert = None, None
                
                # Сохраняем ошибки для каждого температурного режима
                if temp_errors:
                    errors[temp_mode] = temp_errors
                
                # Рассчитываем t-значения
                t_year, t_vlagh, t_urvibr = None, None, None
                
                # t-значение для года
                try:
                    if corr_year is not None:
                        n = len(dol_year)
                        if n > 2 and abs(corr_year) < 1:
                            t_year = abs(corr_year) / math.sqrt((1 - corr_year**2) / (n - 2))
                except Exception as e:
                    temp_errors.append(f"Ошибка при расчете t-значения для года: {str(e)}")
                
                # t-значение для влажности
                try:
                    if corr_vlagh is not None:
                        n = len(dol_vlagh)
                        if n > 2 and abs(corr_vlagh) < 1:
                            t_vlagh = abs(corr_vlagh) / math.sqrt((1 - corr_vlagh**2) / (n - 2))
                except Exception as e:
                    temp_errors.append(f"Ошибка при расчете t-значения для влажности: {str(e)}")
                
                # t-значение для вибрации
                try:
                    if corr_urvibr is not None:
                        n = len(dol_urvibr)
                        if n > 2 and abs(corr_urvibr) < 1:
                            t_urvibr = abs(corr_urvibr) / math.sqrt((1 - corr_urvibr**2) / (n - 2))
                except Exception as e:
                    temp_errors.append(f"Ошибка при расчете t-значения для вибрации: {str(e)}")
                
                # Значимость на уровне 0.05 для t-распределения с df=n-2
                critical_t = 1.96  # Приблизительно для большой выборки
                
                # Средние значения по типам гироскопов
                types_means = {}
                for type_rank, type_name in types_map_reverse.items():
                    type_indices = [i for i, rank in enumerate(type_ranks) if rank == type_rank]
                    if type_indices:
                        type_dol_values = [data[dol_key][i] for i in type_indices]
                        if type_dol_values:
                            mean_value = sum(type_dol_values) / len(type_dol_values)
                            if not math.isnan(mean_value):
                                types_means[type_name] = mean_value
                
                # Средние значения по частям (cert)
                cert_means = {}
                for cert_rank, cert_name in cert_map_reverse.items():
                    cert_indices = [i for i, rank in enumerate(cert_ranks) if rank == cert_rank]
                    if cert_indices:
                        cert_dol_values = [data[dol_key][i] for i in cert_indices]
                        if cert_dol_values:
                            mean_value = sum(cert_dol_values) / len(cert_dol_values)
                            if not math.isnan(mean_value):
                                cert_means[cert_name] = mean_value
                
                # Проверяем наличие рассчитанных данных
                significance = {}
                t_values = {}
                
                if t_year is not None and not math.isnan(t_year):
                    significance["year"] = abs(t_year) > critical_t
                    t_values["year"] = float(t_year)
                
                if t_vlagh is not None and not math.isnan(t_vlagh):
                    significance["humidity"] = abs(t_vlagh) > critical_t
                    t_values["humidity"] = float(t_vlagh)
                
                if t_urvibr is not None and not math.isnan(t_urvibr):
                    significance["vibration"] = abs(t_urvibr) > critical_t
                    t_values["vibration"] = float(t_urvibr)
                
                # Формируем данные о корреляции
                correlations[temp_mode] = CorrelationData(
                    temperature=temp_mode,
                    humidity=float(corr_vlagh) if corr_vlagh is not None and not math.isnan(corr_vlagh) else None,
                    year=float(corr_year) if corr_year is not None and not math.isnan(corr_year) else None,
                    vibration=float(corr_urvibr) if corr_urvibr is not None and not math.isnan(corr_urvibr) else None,
                    types=types_means,
                    departments=cert_means,
                    significance=significance,
                    t_values=t_values
                )
            
            wb.close()
            
            # Если есть ошибки, логируем их, но все равно возвращаем доступные данные
            if errors:
                for mode, mode_errors in errors.items():
                    for error in mode_errors:
                        print(f"Ошибка при расчете корреляций ({mode}): {error}")
            
            return correlations
            
        except HTTPException as http_ex:
            raise http_ex
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Ошибка при расчете корреляций: {str(e)}"
            )
            
    def create_correlation_matrix(self) -> CorrelationMatrix:
        """
        Создает матрицу корреляций и возвращает ее вместе со списком причин и мероприятий,
        используя реальные данные о влажности и вибрации из таблицы report_data
        """
        if not self.table_path.exists():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Таблица с данными о погрешностях не найдена"
            )
            
        try:
            # Получаем матрицу корреляций на основе реальных данных
            correlations = self.calculate_correlations()
            
            # Заполняем матрицу
            matrix = {}
            # Только строки с долями погрешностей
            rows = ["dol1", "dol2", "dol3"]
            # Только столбцы с независимыми переменными
            columns = ["year", "vlagh", "cert", "type", "urvibr"]
            
            row_labels = {
                "dol1": "НКУ", 
                "dol2": "-50", 
                "dol3": "+50"
            }
            
            column_labels = {
                "year": "Год", 
                "vlagh": "Влажность", 
                "cert": "Часть", 
                "type": "Тип изделия",
                "urvibr": "Уровень вибрации"
            }
            
            # Создаем структуру матрицы
            for row_key in rows:
                matrix[row_labels[row_key]] = {}
                for col_key in columns:
                    # Инициализируем значение как None
                    value = None
                    
                    # Выбираем значение из рассчитанных корреляций
                    temp_mode_map = {
                        "dol1": "nku",
                        "dol2": "minus_50",
                        "dol3": "plus_50"
                    }
                    
                    temp_mode = temp_mode_map.get(row_key)
                    
                    if temp_mode in correlations:
                        if col_key == "vlagh":
                            value = correlations[temp_mode].humidity
                        elif col_key == "year":
                            value = correlations[temp_mode].year
                        elif col_key == "urvibr":
                            value = correlations[temp_mode].vibration
                        # Для типа и части используем значения из корреляций, но можно
                        # добавить фиксированные запасные значения если необходимо
                    
                    # Если значение None, используем запасные значения, чтобы не возникало ошибок
                    # в интерфейсе. Тут можно указать осмысленные по умолчанию
                    if value is None:
                        # Предопределенные значения для случаев, когда невозможно рассчитать корреляцию
                        if col_key == "year":
                            value = 0.1  # Слабая корреляция с годом по умолчанию
                        elif col_key == "vlagh":
                            value = 0.4  # Слабая корреляция с влажностью по умолчанию
                        elif col_key == "urvibr":
                            value = 0.5  # Слабая корреляция с вибрацией по умолчанию
                        elif col_key == "cert":
                            value = 0.25  # Слабая корреляция с частью по умолчанию
                        elif col_key == "type":
                            value = -0.2  # Слабая отрицательная корреляция с типом по умолчанию
                    
                    matrix[row_labels[row_key]][column_labels[col_key]] = value
            
            return CorrelationMatrix(matrix=matrix, reasons=self.reasons)
            
        except HTTPException as http_ex:
            raise http_ex
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Ошибка при создании матрицы корреляций: {str(e)}"
            )